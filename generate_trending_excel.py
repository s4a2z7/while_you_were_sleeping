"""
화제 종목 TOP 5 엑셀 데이터 생성
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# 경로 설정
OUTPUT_DIR = "output/data"
os.makedirs(OUTPUT_DIR, exist_ok=True)
OUTPUT_PATH = os.path.join(OUTPUT_DIR, "trending_2025-12-05.xlsx")

# 데이터
data = [
    {"rank": 1, "ticker": "TSLA", "name": "Tesla", "price": "$385.20", "change": "+8.7%"},
    {"rank": 2, "ticker": "NVDA", "name": "NVIDIA", "price": "$142.50", "change": "+5.2%"},
    {"rank": 3, "ticker": "AAPL", "name": "Apple", "price": "$195.80", "change": "+2.1%"},
    {"rank": 4, "ticker": "MSFT", "name": "Microsoft", "price": "$423.50", "change": "-1.3%"},
    {"rank": 5, "ticker": "GOOGL", "name": "Google", "price": "$178.25", "change": "-0.5%"},
]

# 워크북 생성
wb = Workbook()
ws = wb.active
ws.title = "화제 종목"

# 헤더 설정
headers = ["순위", "티커", "종목명", "주가", "등락률"]
ws.append(headers)

# 헤더 스타일
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = border

# 데이터 입력
for row_idx, item in enumerate(data, start=2):
    ws[f"A{row_idx}"] = item["rank"]
    ws[f"B{row_idx}"] = item["ticker"]
    ws[f"C{row_idx}"] = item["name"]
    ws[f"D{row_idx}"] = item["price"]
    ws[f"E{row_idx}"] = item["change"]
    
    # 행 스타일
    for col_idx in range(1, 6):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(size=11)
        
        # 등락률 열 색상 지정
        if col_idx == 5:  # 등락률 열
            if "+" in item["change"]:
                # 상승: 녹색
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                cell.font = Font(size=11, color="006100", bold=True)
            elif "-" in item["change"]:
                # 하락: 빨간색
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                cell.font = Font(size=11, color="9C0006", bold=True)

# 열 너비 설정
ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 10
ws.column_dimensions["C"].width = 15
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 12

# 행 높이 설정
ws.row_dimensions[1].height = 25
for row_idx in range(2, 7):
    ws.row_dimensions[row_idx].height = 22

# 엑셀 파일 저장
wb.save(OUTPUT_PATH)
print(f"✅ 엑셀 파일 생성 완료: {OUTPUT_PATH}")
print(f"📊 데이터 요약:")
print(f"  - 파일명: trending_2025-12-05.xlsx")
print(f"  - 총 5개 종목 데이터")
print(f"  - 상승: 초록색, 하락: 빨간색으로 표시")
print(f"\n📈 포함 종목:")
for item in data:
    print(f"  {item['rank']}. {item['ticker']} ({item['name']}) - {item['change']}")
